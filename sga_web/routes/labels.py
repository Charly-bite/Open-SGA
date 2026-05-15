"""
Label generation routes for SGA Web
"""

import os
import io
import base64
import tempfile
import logging
import traceback
from datetime import date, datetime, timedelta
from flask import (
    Blueprint,
    render_template,
    request,
    jsonify,
    send_file,
    current_app,
    flash,
    redirect,
    url_for,
    session,
)
from flask_login import login_required, current_user

labels_bp = Blueprint("labels", __name__)


def _save_error_payload(smart_label, default_message):
    """Build a user-facing payload with DB path and save error details when available."""
    if hasattr(smart_label, "get_last_save_status"):
        status = smart_label.get_last_save_status() or {}
        if status.get("error"):
            save_path = status.get("path") or getattr(smart_label, "data_dir", "")
            return {
                "error": f"{default_message}. Ruta: {save_path}. Detalle: {status['error']}",
                "save_error": status.get("error"),
                "save_path": save_path,
                "db_mode": status.get("mode")
                or getattr(smart_label, "connection_mode", None),
            }
    return {"error": default_message}


import pandas as pd
from order_status_manager import OrderStatus


def _auto_update_order_status(order_id, username):
    """
    Automatically update the order status to 'En Proceso' when labels are
    printed from the warehouse.  Only promotes orders that are currently
    'Pendiente' so we never overwrite a more advanced state.

    Args:
        order_id: The SAP order number (DocNum).  May be None when printing
                  labels that are not associated with a specific order.
        username: The warehouse user who triggered the print.

    Returns:
        True if the status was updated, False otherwise.
    """
    if not order_id:
        return False

    try:
        order_mgr = current_app.order_status_mgr
        order = order_mgr.get_order(str(order_id))

        if not order:
            logging.info(
                f"Order {order_id} not found in status DB — skipping auto-update"
            )
            return False

        current_status = order.get("status", "")
        if current_status == OrderStatus.PENDING.value:
            order_mgr.update_status(
                str(order_id),
                OrderStatus.IN_PROGRESS.value,
                username,
                "Etiquetas impresas — estado actualizado automáticamente",
            )
            logging.info(f"✅ Order {order_id} auto-updated to 'En Proceso'")
            return True
        else:
            logging.info(
                f"Order {order_id} status is '{current_status}' — no auto-update needed"
            )
            return False
    except Exception as e:
        logging.warning(f"Could not auto-update order {order_id} status: {e}")
        return False


def _get_latest_batch_override(tara_mgr, history_mgr, product_code):
    """
    Get the dynamically computed Lote override from Control Interno's history
    to fix labels dragging old batch numbers when SAP overwrites them.
    """
    try:
        from routes.control_interno import _extract_user_lote_overrides

        overrides = _extract_user_lote_overrides(tara_mgr, history_mgr)
        return overrides.get(product_code, {})
    except Exception as e:
        import logging

        logging.error(f"Error loading overrides: {e}")
        return {}


def get_batch_and_tare(product_code, quantity):
    """
    Get (peso_tara, batch_number, batch_date, reinspection_date) from TaraWeightManager.
    """
    peso_tara = 0.0
    lote = ""
    fecha = ""
    vencimiento = ""

    try:
        tara_mgr = current_app.tara_manager
        qty = float(quantity) if quantity else 0
        peso_tara = tara_mgr.resolve_tara(product_code, qty)

        classification = tara_mgr.get_classification(product_code)
        if classification:
            lote = classification.get("lote", "") or classification.get(
                "batch_number", ""
            )
            if lote and "," in str(lote):
                lote = str(lote).split(",")[-1].strip()
            fecha = classification.get("lote_date", "")
            vencimiento = classification.get("lote_reinspection_date", "")

            # P1-03: FIX - "the labels still drag the old batch number"
            # Recover the latest user modification from Control Interno history
            history_mgr = getattr(current_app, "history_mgr", None)
            if history_mgr:
                override = _get_latest_batch_override(
                    tara_mgr, history_mgr, product_code
                )
                if override:
                    lote = override.get("lote", lote)
                    fecha = override.get("lote_date", fecha)
                    vencimiento = override.get("lote_reinspection_date", vencimiento)
    except Exception as e:
        logging.error(f"Error getting batch and tare for {product_code}: {e}")

    return peso_tara, lote, fecha, vencimiento


def _load_template(template_id):
    """
    Load a template dict by ID via the app's TemplateManager.
    Returns None if template_id is falsy, 'default', or not found.
    """
    if not template_id or template_id == "default":
        return None
    try:
        return current_app.template_manager.get_template(template_id)
    except Exception as e:
        logging.warning(f"Could not load template '{template_id}': {e}")
        return None


def _generate_label(generator, product_data, output_path, template=None):
    """
    Generate a label PDF.
    Uses `generate_from_template` when a template dict is provided,
    falls back to the standard `generate_label` otherwise.
    Returns (page_w_mm, page_h_mm) based on the template dimensions or
    warehouse-based heuristic.
    """
    if template:
        generator.generate_from_template(product_data, template, output_path)
        return template.get("width_mm", 150), template.get("height_mm", 100)
    else:
        generator.generate_label(product_data, output_path)
        wh = str(product_data.get("warehouse", "")).strip()
        if wh in ("02", "2"):
            return 200, 150
        return 150, 100


# ---------------------------------------------------------------------------
# Shared helpers (CODE-01 deduplication)
# ---------------------------------------------------------------------------


def _get_poppler_path():
    """Return cached poppler path from app config (resolved once at startup)."""
    return current_app.config.get("POPPLER_PATH")


def _build_product_data(item, use_m_y_format=False):
    """Merge queue-item overrides into a copy of the resolved product data.
    If the queue item only stores a product code (PERF-05 optimisation),
    we re-resolve the product data from SmartLabelManager."""
    product_data = item.get("product_data")
    if product_data:
        product_data = product_data.copy()
    else:
        # Re-resolve from SmartLabelManager (session no longer stores full dict)
        code = item.get("code", "")
        product_data = current_app.smart_label.get_product_data(code)
        if not product_data:
            product_data = {"product_id": code}
        else:
            product_data = product_data.copy()

    # Apply queue-specific overrides
    product_data["batch_number"] = item.get("batch_number", "000000")
    product_data["batch_date"] = item.get("batch_date", "")
    product_data["reinspection_date"] = item.get("reinspection_date", "")
    product_data["peso_tara"] = item.get("peso_tara", 0)
    product_data["quantity"] = item.get("quantity", 0)
    product_data["warehouse"] = item.get("warehouse", "")
    product_data["use_m_y_format"] = use_m_y_format
    return product_data


def _resolve_queue_items(item_ids=None):
    """Return (items_to_generate, full_queue) from session.
    If item_ids is given, filter; otherwise return entire queue."""
    queue = session.get("print_queue", [])
    if item_ids:
        ids_int = [int(i) for i in item_ids]
        items = [i for i in queue if i.get("id") in ids_int]
    else:
        items = []
    if not items:
        items = list(queue)  # fallback: all
    return items, queue


def _determine_page_size(items, template):
    """Return (width_mm, height_mm) based on template or warehouse heuristic."""
    if template:
        return template.get("width_mm", 150), template.get("height_mm", 100)
    first_wh = str(items[0].get("warehouse", "")).strip() if items else ""
    if first_wh in ("02", "2"):
        return 200, 150
    return 150, 100


@labels_bp.route("/")
@login_required
def index():
    """Label generation main page"""
    user_warehouse = getattr(current_user, "warehouse", "") or ""
    return render_template("labels/index.html", user_warehouse=user_warehouse)


@labels_bp.route("/api/quick-add-product", methods=["POST"])
@login_required
def quick_add_product():
    """Quick-add a product from the labels page — available to Operators and Admins"""
    if not current_user.can_add_products():
        return jsonify({"error": "Sin permisos para agregar productos"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"error": "No se recibieron datos"}), 400

    # Validate required fields
    code = (data.get("Codigo interno") or "").strip()
    name = (data.get("Chemical Name") or "").strip()
    if not code or not name:
        return jsonify({"error": "Código interno y Nombre son requeridos"}), 400

    smart_label = current_app.smart_label

    # Check if already exists
    if smart_label.get_product_data(code):
        return (
            jsonify({"error": f"El código {code} ya existe en la base de datos"}),
            400,
        )

    # Add to database
    success = smart_label.add_product(data)
    if not success:
        payload = _save_error_payload(smart_label, "Error al agregar producto")
        logging.warning(f"Quick-add save failed for {code}: {payload}")
        return jsonify(payload), 500

    # Log
    try:
        current_app.history_mgr.add_entry(
            event_type="product_add",
            username=current_user.username,
            details={"product_code": code, "source": "quick_add"},
        )
    except Exception:
        pass

    # Refresh label generator cache
    try:
        from generate_ghs_label import GHSLabelGenerator

        current_app.label_generator = GHSLabelGenerator(
            current_app.config["UNIFIED_DB_PATH"], manager=current_app.smart_label
        )
    except Exception as e:
        logging.warning(f"Could not refresh label generator after quick-add: {e}")

    return jsonify({"success": True, "code": code, "name": name})


@labels_bp.route("/queue", methods=["GET"])
@login_required
def get_queue():
    """Get current print queue from session or temp storage"""
    queue = session.get("print_queue", [])
    return jsonify({"queue": queue})


@labels_bp.route("/queue/add", methods=["POST"])
@login_required
def add_to_queue():
    """Add item to print queue"""
    if not current_user.can_print_labels():
        return jsonify({"error": "Sin permisos para imprimir etiquetas"}), 403

    data = request.get_json()
    if not data:
        logging.warning("No JSON data received in request")
        return jsonify({"error": "No se recibieron datos"}), 400

    code = data.get("code")
    if code:
        code = str(code).strip()
    quantity = data.get("quantity", 1)
    peso_tara = data.get("peso_tara", 0)  # Get tare weight from request
    batch_number = data.get("batch_number", "")  # Get batch number from request
    batch_date = data.get("batch_date", "")  # Get batch date from request

    if batch_date:
        try:
            parsed_bd = datetime.strptime(batch_date[:10], "%Y-%m-%d")
            if parsed_bd.date() > datetime.now().date():
                return (
                    jsonify(
                        {
                            "error": f"La fecha del lote ({batch_date}) no puede estar en el futuro."
                        }
                    ),
                    400,
                )
        except ValueError:
            pass

    reinspection_date = data.get(
        "reinspection_date", ""
    )  # Get reinspection date from request
    copies = data.get("copies", 1)  # Number of copies to print
    warehouse = data.get("warehouse", "")  # Get warehouse from request

    if not code:
        logging.warning(f"Empty code received. Full data: {data}")
        return jsonify({"error": "Código requerido"}), 400

    # Resolve product using SmartLabelManager
    smart_label = current_app.smart_label
    product_data = smart_label.get_product_data(code)

    if not product_data:
        logging.warning(f"Product not found: {code}")
        return jsonify({"error": f"Producto no encontrado: {code}"}), 404

    # Create queue item
    # SmartLabelManager returns 'name' or 'product_name', not 'Chemical Name'
    product_name = (
        product_data.get("product_name")
        or product_data.get("name")
        or product_data.get("Chemical Name", "Unknown")
    )

    # Generate unique ID based on max existing ID
    queue = session.get("print_queue", [])
    if queue:
        new_id = max((item.get("id", 0) for item in queue), default=0) + 1
    else:
        new_id = 1

    # Lookup Batch and Tare (Prioritize Control Interno database over SAP HANA parsed payload)
    db_tara, db_batch, db_date, db_vencimiento = get_batch_and_tare(code, quantity)
    if hasattr(
        db_batch, "values"
    ):  # Handle if pandas series returned by mistake (safety)
        db_batch = str(db_batch.iloc[0])

    if not batch_number and db_batch:
        batch_number = db_batch

    if not batch_date and db_date:
        batch_date = db_date
    if not reinspection_date and db_vencimiento:
        reinspection_date = db_vencimiento

    # Only override tare if it wasn't provided and we found one
    if peso_tara == 0:
        peso_tara = db_tara

    # ── TaraWeightManager smart fallback ──────────────────────────────
    # If tara is still 0 after SAP UDF + CSV lookup, use the smart
    # suggestion engine (statistical + product history interpolation).
    if peso_tara == 0 and quantity and float(quantity) > 0:
        try:
            tara_mgr = current_app.tara_manager
            peso_tara = tara_mgr.resolve_best_tara(float(quantity), code)
        except Exception as e:
            logging.warning(
                f"TaraWeightManager fallback failed for {code}/{quantity}: {e}"
            )

    # Auto-calculate reinspection date if not provided (batch_date + 1 year)
    if not reinspection_date and batch_date:
        try:
            from datetime import timedelta

            bd = datetime.strptime(batch_date, "%Y-%m-%d")
            try:
                rd = bd.replace(year=bd.year + 1)
            except ValueError:
                rd = bd.replace(year=bd.year + 1, day=28)
            reinspection_date = rd.strftime("%Y-%m-%d")
        except Exception:
            reinspection_date = ""

    item = {
        "id": new_id,
        "code": code,
        "product_name": product_name,
        "product_type": current_app.tara_manager.get_classification(code).get(
            "product_type", ""
        ),
        "quantity": quantity,
        "peso_tara": peso_tara,
        "peso_bruto": 0,
        "batch_number": batch_number if batch_number else "",
        "batch_date": batch_date if batch_date else date.today().isoformat(),
        "reinspection_date": reinspection_date if reinspection_date else "",
        "copies": max(1, int(copies)) if copies else 1,
        "warehouse": warehouse,
        # PERF-05: product_data no longer stored in session to avoid
        # 4 KB cookie overflow.  Resolved on-the-fly via _build_product_data().
    }

    # Add to session queue
    queue = session.get("print_queue", [])
    queue.append(item)
    session["print_queue"] = queue

    return jsonify({"success": True, "item": item, "queue_size": len(queue)})


@labels_bp.route("/queue/remove/<int:item_id>", methods=["DELETE"])
@login_required
def remove_from_queue(item_id):
    """Remove item from print queue"""
    queue = session.get("print_queue", [])
    queue = [item for item in queue if item.get("id") != item_id]
    session["print_queue"] = queue

    return jsonify({"success": True, "queue_size": len(queue)})


@labels_bp.route("/queue/clear", methods=["POST"])
@login_required
def clear_queue():
    """Clear entire print queue"""
    session["print_queue"] = []
    session.modified = True
    return jsonify({"success": True})


@labels_bp.route("/queue/update/<int:item_id>", methods=["PUT"])
@login_required
def update_queue_item(item_id):
    """Update queue item (batch info, quantity, etc.)"""
    data = request.get_json()
    queue = session.get("print_queue", [])

    # Validate batch_date if updated
    new_batch_date = data.get("batch_date", "")
    if new_batch_date:
        try:
            parsed_bd = datetime.strptime(new_batch_date[:10], "%Y-%m-%d")
            if parsed_bd.date() > datetime.now().date():
                return (
                    jsonify(
                        {
                            "error": f"La fecha del lote ({new_batch_date}) no puede estar en el futuro."
                        }
                    ),
                    400,
                )
        except ValueError:
            pass

    for item in queue:
        if item.get("id") == item_id:
            item["batch_number"] = data.get(
                "batch_number", item.get("batch_number", "")
            )
            item["batch_date"] = data.get("batch_date", item.get("batch_date", ""))
            item["reinspection_date"] = data.get(
                "reinspection_date", item.get("reinspection_date", "")
            )
            item["copies"] = max(1, int(data.get("copies", item.get("copies", 1))))
            item["quantity"] = data.get("quantity", item.get("quantity", 1))
            item["peso_tara"] = data.get("peso_tara", item.get("peso_tara", 0))
            item["peso_bruto"] = data.get("peso_bruto", item.get("peso_bruto", 0))
            break

    session["print_queue"] = queue
    return jsonify({"success": True})


@labels_bp.route("/preview/<int:item_id>")
@login_required
def preview_label(item_id):
    """Generate and return label preview"""
    queue = session.get("print_queue", [])
    item = next((i for i in queue if i.get("id") == item_id), None)

    if not item:
        return jsonify({"error": "Item not found"}), 404

    try:
        generator = current_app.label_generator

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            output_path = tmp.name

        product_data = _build_product_data(item)

        template_id = request.args.get("template_id")
        template = _load_template(template_id)
        _generate_label(generator, product_data, output_path, template)

        response = send_file(
            output_path, mimetype="application/pdf", as_attachment=False
        )
        response.headers["Cache-Control"] = (
            "no-store, no-cache, must-revalidate, max-age=0"
        )
        return response

    except Exception as e:
        logging.error(f"Error in preview_label: {e}")
        return jsonify({"error": str(e)}), 500


@labels_bp.route("/generate", methods=["POST"])
@login_required
def generate_labels():
    """Generate labels for selected items"""
    import time
    from generate_ghs_label import GHSLabelGenerator
    import zipfile

    if not current_user.can_print_labels():
        return jsonify({"error": "Sin permisos para generar etiquetas"}), 403

    start_time = time.perf_counter()
    data = request.get_json()
    item_ids = data.get("item_ids", [])
    template_id = data.get("template_id")
    use_m_y_format = data.get("use_m_y_format", False)

    items_to_generate, queue = _resolve_queue_items(item_ids)
    logging.info(
        f"Generate labels: queue has {len(queue)} items, requested ids: {item_ids}, template: {template_id}"
    )

    if not items_to_generate:
        return jsonify({"error": "No hay items en la cola"}), 400

    logging.info(f"Will generate {len(items_to_generate)} labels")

    try:
        # Use cached generator (loaded once at startup)
        generator = current_app.label_generator
        template = _load_template(template_id)
        generated_files = []  # Temp files for individual labels

        for item in items_to_generate:
            tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            tmp_path = tmp.name
            tmp.close()

            logging.info(f"Generating label for {item['code']} -> {tmp_path}")

            product_data = _build_product_data(item, use_m_y_format=use_m_y_format)

            _generate_label(generator, product_data, tmp_path, template)
            # Add file multiple times for copies
            num_copies = max(1, int(item.get("copies", 1)))
            for _ in range(num_copies):
                generated_files.append(tmp_path)
            logging.info(f"Label generated: {tmp_path} (x{num_copies} copies)")

        # Merge all individual PDFs into a single multi-page PDF using PyMuPDF (fitz)
        import fitz

        output_dir = current_app.config["GENERATED_LABELS_PATH"]
        from datetime import datetime

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        final_pdf_path = os.path.join(
            output_dir, f"etiquetas_{timestamp}_{len(generated_files)}pcs.pdf"
        )

        merged_pdf = fitz.open()
        for f in generated_files:
            try:
                pdf_doc = fitz.open(f)
                merged_pdf.insert_pdf(pdf_doc)
                pdf_doc.close()
            except Exception as e:
                logging.warning(f"Error merging label {f}: {e}")

        merged_pdf.save(final_pdf_path)
        merged_pdf.close()

        logging.info(f"Merged {len(generated_files)} labels into {final_pdf_path}")

        # Clean up individual temp files (use set to avoid duplicate removal)
        for f in set(generated_files):
            try:
                os.remove(f)
            except Exception:
                pass

        # Log to history
        current_app.history_mgr.add_entry(
            event_type="PRINT_JOB",
            username=current_user.username,
            details={
                "count": len(items_to_generate),
                "items": [i["code"] for i in items_to_generate],
            },
        )

        # Clear generated items from queue
        generated_ids = [i.get("id") for i in items_to_generate]
        remaining_queue = [i for i in queue if i.get("id") not in generated_ids]
        session["print_queue"] = remaining_queue

        # Auto-update order status to "En Proceso" when labels are printed
        _auto_update_order_status(data.get("order_id"), current_user.username)

        # Return the merged PDF to the client browser for client-side printing
        # This allows each client to print on their own local printer
        logging.info(f"Returning merged PDF to client for printing: {final_pdf_path}")
        return send_file(
            final_pdf_path,
            mimetype="application/pdf",
            as_attachment=False,
            download_name=f"etiquetas_{timestamp}_{len(generated_files)}pcs.pdf",
        )

    except Exception as e:
        logging.error(f"Error generating labels: {e}")
        logging.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@labels_bp.route("/generate_print", methods=["POST"])
@login_required
def generate_print_page():
    """
    Generate label images and render them into an HTML print queue page.
    This bypasses the native PDF viewer and prevents scaling issues by
    forcing images at exact mm sizes.
    """
    if not current_user.can_print_labels():
        return jsonify({"error": "Sin permisos para generar etiquetas"}), 403

    data = request.get_json()
    item_ids = data.get("item_ids", [])
    template_id = data.get("template_id")
    use_m_y_format = data.get("use_m_y_format", False)
    PRINT_DPI = min(300, int(data.get("print_dpi", 600)))

    items_to_generate, queue = _resolve_queue_items(item_ids)

    if not items_to_generate:
        return jsonify({"error": "No hay items en la cola"}), 400

    try:
        generator = current_app.label_generator
        template = _load_template(template_id)
        page_w_mm, page_h_mm = _determine_page_size(items_to_generate, template)

        poppler_path = _get_poppler_path()

        from pdf2image import convert_from_path
        import concurrent.futures

        orientation = "landscape" if page_w_mm > page_h_mm else "portrait"
        result_images = []
        app_context = current_app._get_current_object()

        def process_item(item_data):
            with app_context.app_context():
                idx, item = item_data
                product_data = _build_product_data(item, use_m_y_format=use_m_y_format)

                tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
                tmp_path = tmp.name
                tmp.close()

                _generate_label(generator, product_data, tmp_path, template)

                num_copies = max(1, int(item.get("copies", 1)))
                local_images = []
                try:
                    images = convert_from_path(
                        tmp_path,
                        dpi=PRINT_DPI,
                        first_page=1,
                        last_page=1,
                        poppler_path=poppler_path,
                    )

                    if images:
                        img = images[0]
                        buf = io.BytesIO()
                        img.save(buf, "PNG", dpi=(PRINT_DPI, PRINT_DPI))
                        img_base64 = base64.b64encode(buf.getvalue()).decode("utf-8")

                        for copy_idx in range(num_copies):
                            local_images.append(
                                {
                                    "data": img_base64,
                                    "width_mm": page_w_mm,
                                    "height_mm": page_h_mm,
                                }
                            )
                except Exception as e:
                    logging.error(f"PDF to image conversion failed for item {idx}: {e}")
                finally:
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass
                return local_images

        # Process labels in parallel to speed up pdftoppm generation
        with concurrent.futures.ThreadPoolExecutor(
            max_workers=os.cpu_count() or 4
        ) as executor:
            for item_images in executor.map(process_item, enumerate(items_to_generate)):
                result_images.extend(item_images)

        if not result_images:
            return jsonify({"error": "No se pudieron generar las imágenes"}), 500

        current_app.history_mgr.add_entry(
            event_type="PRINT_JOB_HTML",
            username=current_user.username,
            details={
                "count": len(items_to_generate),
                "items": [i["code"] for i in items_to_generate],
            },
        )

        generated_ids = [i.get("id") for i in items_to_generate]
        remaining_queue = [i for i in queue if i.get("id") not in generated_ids]
        session["print_queue"] = remaining_queue

        # Auto-update order status to "En Proceso" when labels are printed
        _auto_update_order_status(data.get("order_id"), current_user.username)

        return render_template(
            "labels/print.html",
            images=result_images,
            orientation=orientation,
            page_w_mm=page_w_mm,
            page_h_mm=page_h_mm,
        )

    except Exception as e:
        logging.error(f"Error generating print page: {e}")
        logging.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@labels_bp.route("/generate_images", methods=["POST"])
@login_required
def generate_images_api():
    """
    Generate label images and return them as base64-encoded JSON.
    Used by the local Print Agent to receive image data directly,
    bypassing Chrome's print dialog entirely.

    Request:  { "item_ids": [1, 2, 3], "rotation": 90 }
    Response: { "images": [{ "data": "<base64>", "width_mm": 200, ... }], ... }
    """
    if not current_user.can_print_labels():
        return jsonify({"error": "Sin permisos para generar etiquetas"}), 403

    data = request.get_json()
    item_ids = data.get("item_ids", [])
    template_id = data.get("template_id")
    use_m_y_format = data.get("use_m_y_format", False)
    print_rotation = int(data.get("rotation", 0))
    PRINT_DPI = int(data.get("print_dpi", 600))

    items_to_generate, queue = _resolve_queue_items(item_ids)

    if not items_to_generate:
        return jsonify({"error": "No hay items en la cola"}), 400

    try:
        generator = current_app.label_generator
        template = _load_template(template_id)
        page_w_mm, page_h_mm = _determine_page_size(items_to_generate, template)

        poppler_path = _get_poppler_path()

        from pdf2image import convert_from_path

        result_images = []

        for idx, item in enumerate(items_to_generate):
            product_data = _build_product_data(item, use_m_y_format=use_m_y_format)

            # Generate PDF to temp file
            tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            tmp_path = tmp.name
            tmp.close()

            _generate_label(generator, product_data, tmp_path, template)

            num_copies = max(1, int(item.get("copies", 1)))
            try:
                images = convert_from_path(
                    tmp_path,
                    dpi=PRINT_DPI,
                    first_page=1,
                    last_page=1,
                    poppler_path=poppler_path,
                )

                if images:
                    img = images[0]

                    # Apply additional rotation for direct print (90° CW = -90° in PIL)
                    rotated_w, rotated_h = page_w_mm, page_h_mm
                    if print_rotation == 90:
                        img = img.rotate(-90, expand=True)
                        rotated_w, rotated_h = page_h_mm, page_w_mm
                    elif print_rotation == 180:
                        img = img.rotate(180, expand=True)
                    elif print_rotation == 270:
                        img = img.rotate(90, expand=True)
                        rotated_w, rotated_h = page_h_mm, page_w_mm

                    buf = io.BytesIO()
                    img.save(buf, "PNG", dpi=(PRINT_DPI, PRINT_DPI))
                    img_base64 = base64.b64encode(buf.getvalue()).decode("utf-8")

                    # Add image multiple times for copies
                    for copy_idx in range(num_copies):
                        result_images.append(
                            {
                                "data": img_base64,
                                "width_mm": rotated_w,
                                "height_mm": rotated_h,
                                "code": item.get("code", ""),
                                "name": item.get("name", ""),
                                "index": idx,
                                "copy": copy_idx + 1,
                            }
                        )
            except Exception as e:
                logging.error(f"PDF to image conversion failed for item {idx}: {e}")
            finally:
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

        if not result_images:
            return jsonify({"error": "No se pudieron generar las imágenes"}), 500

        # Log to history
        current_app.history_mgr.add_entry(
            event_type="DIRECT_PRINT_JOB",
            username=current_user.username,
            details={
                "count": len(items_to_generate),
                "items": [i["code"] for i in items_to_generate],
                "method": "print_agent",
            },
        )

        # Clear printed items from queue
        generated_ids = [i.get("id") for i in items_to_generate]
        remaining_queue = [i for i in queue if i.get("id") not in generated_ids]
        session["print_queue"] = remaining_queue

        # Auto-update order status to "En Proceso" when labels are printed
        order_status_updated = _auto_update_order_status(
            data.get("order_id"), current_user.username
        )

        # Determine final page dimensions (account for rotation)
        final_w, final_h = page_w_mm, page_h_mm
        if print_rotation in (90, 270):
            final_w, final_h = page_h_mm, page_w_mm

        return jsonify(
            {
                "success": True,
                "images": result_images,
                "page_width_mm": final_w,
                "page_height_mm": final_h,
                "count": len(result_images),
                "order_status_updated": order_status_updated,
            }
        )

    except Exception as e:
        logging.error(f"Error generating images for print agent: {e}")
        logging.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@labels_bp.route("/sap/load", methods=["POST"])
@login_required
def load_from_sap():
    """Load order items from SAP"""
    if not current_app.sap_available:
        return jsonify({"error": "SAP no disponible"}), 503

    data = request.get_json()
    order_number = data.get("order_number", "").strip()

    if not order_number:
        return jsonify({"error": "Número de pedido requerido"}), 400

    # ARCH-01: Validate order number format (numeric only, max 10 digits)
    if not order_number.isdigit() or len(order_number) > 10:
        return jsonify({"error": "Número de pedido inválido"}), 400

    try:
        # Use pre-connected SAP or try to connect
        sap = current_app.sap_connector
        if not sap:
            return (
                jsonify(
                    {"error": "SAP no conectado. Configure las credenciales en .env"}
                ),
                503,
            )

        if not sap.connected:
            sap.connect()

        # Get order details
        order_data = sap.get_order_details(order_number)

        if not order_data:
            return jsonify({"error": f"Pedido {order_number} no encontrado"}), 404

        # Transform SAP response to expected format for frontend
        # SAP returns: {'header': {...}, 'items': [{'item_code': ...}]}
        # Frontend expects: {'DocNum': ..., 'CardName': ..., 'items': [{'ItemCode': ...}]}
        header = order_data.get("header", {})
        items = order_data.get("items", [])

        enriched_items = []
        for item in items:
            item_code = item.get("item_code")
            quantity = item.get("quantity", 0)

            # -------------------------------------------------------------------
            # P1-02: Use SAP Order Line UDFs as primary source for tare weight
            # -------------------------------------------------------------------
            u_tara = item.get("u_tara", 0.0)  # direct from RDR1
            u_num_etiqueta = item.get("u_num_etiqueta", 0)  # label count
            u_presentacion = item.get("u_presentacion", "")  # packaging type
            u_kilos_pre = item.get("u_kilos_pre", 0.0)  # pre-calculated kg

            # Siempre obtenemos el lote mas reciente de la BD de control interno (Legacy CSV)
            db_peso_tara, lote, fecha_lote, fecha_vencimiento = get_batch_and_tare(
                item_code, quantity
            )

            # Start with SAP UDF tare; fall back to local CSV only if 0
            if u_tara > 0:
                peso_tara = u_tara
            else:
                peso_tara = db_peso_tara

            # If still 0, use SAP OITM BWeight2 (T-02: refactored to sap_connector method)
            if peso_tara == 0:
                try:
                    weights = sap.get_item_weights(item_code)
                    peso_tara = weights.get("tare_weight", 0.0)
                except Exception as e:
                    logging.warning(
                        f"Could not fetch tare weight from SAP for {item_code}: {e}"
                    )

            # ── TaraWeightManager smart fallback ──────────────────────────
            if peso_tara == 0 and quantity and float(quantity) > 0:
                try:
                    tara_mgr = current_app.tara_manager
                    peso_tara = tara_mgr.resolve_best_tara(float(quantity), item_code)
                    if peso_tara > 0:
                        logging.info(
                            f"✅ TaraWeightManager resolved tara for {item_code} @ {quantity}kg → {peso_tara} kg"
                        )
                except Exception as e:
                    logging.warning(
                        f"TaraWeightManager fallback failed for {item_code}/{quantity}: {e}"
                    )

            # -------------------------------------------------------------------
            # P1-01: Try SAP GHS UDF data first; fall back to local CSV (via
            #        SmartLabelManager which is already loaded in the app)
            # -------------------------------------------------------------------
            ghs_from_sap = None
            try:
                ghs_from_sap = sap.get_item_ghs_data(item_code)
            except Exception as e:
                logging.warning(
                    f"Could not fetch GHS UDF data from SAP for {item_code}: {e}"
                )

            # -------------------------------------------------------------------
            # Batch extraction from SAP has been decoupled here per ARCH P1-O5.
            # We strictly prioritize the internal database `fecha_vencimiento`.
            # -------------------------------------------------------------------
            # -------------------------------------------------------------------
            # Fallback: if no batch manufacturing date found, use order DocDate
            # instead of letting the system default to today's date.
            # The user can still manually edit the date in the UI before printing.
            # -------------------------------------------------------------------
            if not fecha_lote:
                order_date = header.get("order_date", "")
                if order_date:
                    fecha_lote = order_date
                    logging.info(
                        f"📅 No batch date for {item_code}, using order date: {order_date}"
                    )

            enriched_item = {
                "ItemCode": item_code,
                "Description": item.get("description"),
                "Quantity": quantity,
                "Unit": item.get("unit"),
                "UnitPrice": item.get("unit_price"),
                "LineTotal": item.get("line_total"),
                "Warehouse": item.get("warehouse"),
                "PesoTara": peso_tara,
                "Lote": lote,
                "FechaLote": fecha_lote,
                "FechaVencimiento": fecha_vencimiento,
                # P1-02 UDF extras
                "NumEtiquetas": u_num_etiqueta,
                "Presentacion": u_presentacion,
                "KilosPre": u_kilos_pre,
            }

            # P1-01: Attach SAP GHS data if found (frontend can use or override from local CSV)
            if ghs_from_sap:
                enriched_item["ghs_sap"] = ghs_from_sap

            enriched_items.append(enriched_item)

        import math

        transformed_order = {
            "DocNum": header.get("order_number"),
            "CardCode": header.get("customer_code"),
            "CardName": header.get("customer_name"),
            "DocDate": header.get("order_date"),
            "DocDueDate": header.get("delivery_date"),
            "DocTotal": header.get("total_value", 0),
            "DocCurrency": header.get("currency", "MXN"),
            "sap_status": header.get("sap_status", "Abierto"),
            "items": enriched_items,
        }

        # Validate against NaN serialization breaking frontend JSON parser
        import math
        import pandas as pd

        def sanitize_nans(obj):
            if isinstance(obj, dict):
                return {k: sanitize_nans(v) for k, v in obj.items()}
            elif isinstance(obj, (list, tuple)):
                return [sanitize_nans(v) for v in obj]
            try:
                if pd.isna(obj):
                    return None
                elif isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
                    return None
            except Exception:
                pass
            return obj

        transformed_order = sanitize_nans(transformed_order)

        # Return order info
        return jsonify({"success": True, "order": transformed_order})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@labels_bp.route("/direct/<code_or_id>")
@login_required
def direct_print(code_or_id):
    """Generate and return label directly, bypassing queue"""
    try:
        # Resolve product using SmartLabelManager
        smart_label = current_app.smart_label
        product_data = smart_label.get_product_data(code_or_id)

        if not product_data:
            return jsonify({"error": f"Producto no encontrado: {code_or_id}"}), 404

        # Use cached generator
        generator = current_app.label_generator

        # Create temp file
        import tempfile

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            output_path = tmp.name

        # Resolve best guess for batch/tare (default quantity 1)
        # In direct print, we don't have user input for batch/tare, so we use defaults
        peso_tara, lote, fecha = get_batch_and_tare(code_or_id, 1)

        if hasattr(lote, "values"):
            lote = str(lote.iloc[0])

        # Prepare data for generator
        label_data = product_data.copy()
        label_data.update(
            {
                "batch_number": lote if lote else "000000",
                "batch_date": fecha,
                "peso_tara": peso_tara,
                "quantity": 1,
            }
        )

        generator.generate_label(label_data, output_path)

        # Log to history
        current_app.history_mgr.add_entry(
            event_type="PRINT_DIRECT",
            username=current_user.username,
            details={"product": code_or_id},
        )

        response = send_file(
            output_path, mimetype="application/pdf", as_attachment=False
        )
        response.headers["Cache-Control"] = (
            "no-store, no-cache, must-revalidate, max-age=0"
        )
        return response

    except Exception as e:
        logging.error(f"Error in direct print: {e}")
        logging.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@labels_bp.route("/print_order/<order_id>")
@login_required
def print_order_labels(order_id):
    """
    Generate labels for an entire order and update status to 'Preparando'.
    """
    try:
        # 1. Get order details
        order_mgr = current_app.order_status_mgr
        order = order_mgr.get_order(order_id)

        if not order:
            return f"Pedido {order_id} no encontrado", 404

        items = order.get("items", [])
        if not items:
            return f"El pedido {order_id} no tiene productos", 400

        # 2. Generate labels
        generator = current_app.label_generator
        generated_files = []

        for item in items:
            item_code = item.get("ItemCode", item.get("item_code"))
            if not item_code:
                continue

            quantity = float(item.get("Quantity", item.get("quantity", 1)))

            # Get batch/tare info
            peso_tara, lote, fecha = get_batch_and_tare(item_code, quantity)
            if hasattr(lote, "values"):
                lote = str(lote.iloc[0])

            # Create temp file
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp_path = tmp.name

            # Prepare data
            # Improve product data fetching if possible, otherwise rely on code
            smart_label = current_app.smart_label
            product_data = smart_label.get_product_data(item_code) or {
                "product_id": item_code
            }

            label_data = product_data.copy()
            label_data.update(
                {
                    "batch_number": lote if lote else "000000",
                    "batch_date": fecha,
                    "peso_tara": peso_tara,
                    "quantity": quantity,
                    "warehouse": item.get("Warehouse", ""),
                }
            )

            generator.generate_label(label_data, tmp_path)
            generated_files.append(tmp_path)
            logging.info(f"Generated label for {item_code} in order {order_id}")

        if not generated_files:
            return "No se pudieron generar etiquetas", 500

        # 3. Merge PDFs
        import fitz

        output_dir = current_app.config["GENERATED_LABELS_PATH"]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        final_pdf_path = os.path.join(output_dir, f"pedido_{order_id}_{timestamp}.pdf")

        merged_pdf = fitz.open()
        for f in generated_files:
            try:
                pdf_doc = fitz.open(f)
                merged_pdf.insert_pdf(pdf_doc)
                pdf_doc.close()
            except Exception as e:
                logging.error(f"Error merging file {f}: {e}")

        merged_pdf.save(final_pdf_path)
        merged_pdf.close()

        # Cleanup
        for f in generated_files:
            try:
                os.remove(f)
            except Exception:
                pass

        # 4. Update Status to 'En Proceso' when labels are printed
        # Only if current status is Pending/Open to avoid overwriting advanced states
        current_status = order.get("status")
        if current_status in [OrderStatus.PENDING.value, "Abierto"]:
            order_mgr.update_status(
                order_id,
                OrderStatus.IN_PROGRESS.value,
                current_user.username,
                "Etiquetas impresas - Pedido en proceso",
            )

        # Log history
        current_app.history_mgr.add_entry(
            event_type="PRINT_ORDER",
            username=current_user.username,
            details={"order_id": order_id, "items": len(generated_files)},
        )

        end_time = time.perf_counter()
        generation_time = end_time - start_time

        # Log performance telemetry
        current_app.history_mgr.add_entry(
            event_type="LABEL_GENERATION_METRICS",
            username=current_user.username,
            details={
                "duration_seconds": round(generation_time, 3),
                "total_copies": len(generated_files),
                "template_id": template_id,
                "order_id": order_id,
            },
        )

        # 5. Return PDF
        response = send_file(
            final_pdf_path,
            mimetype="application/pdf",
            as_attachment=False,
            download_name=f"pedido_{order_id}.pdf",
        )
        response.headers["Cache-Control"] = (
            "no-store, no-cache, must-revalidate, max-age=0"
        )
        return response

    except Exception as e:
        logging.error(f"Error printing order {order_id}: {e}")
        logging.error(traceback.format_exc())
        return f"Error imprimiendo pedido: {str(e)}", 500


@labels_bp.route("/export/lotes_modificados", methods=["GET"])
@login_required
def export_lotes_modificados():
    """Export modified lotes report as an Excel file.

    Pulls from the product classifications all the lote changes
    and generates an .xlsx workbook, replacing the old mermas report.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tara_mgr = current_app.tara_manager
    history_list = []

    for pid, data in tara_mgr._product_classifications.items():
        lh = data.get("lote_history", [])
        for entry in lh:
            enriched = dict(entry)
            enriched["product_id"] = pid
            enriched["chemical_name"] = data.get("chemical_name", "")
            history_list.append(enriched)

    # Filter by month if provided (e.g. YYYY-MM)
    filter_month = request.args.get("month", "")
    if filter_month:
        history_list = [
            h for h in history_list if str(h.get("date", "")).startswith(filter_month)
        ]

    # Sort by descending date
    history_list.sort(key=lambda x: x.get("date", ""), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Lotes Modificados"

    # ── Styles ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="2563EB", end_color="2563EB", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Código Producto",
        "Producto",
        "Fecha de Cambio",
        "Usuario Modificó",
        "Lote Anterior",
        "Lote Nuevo",
        "Fabricación Ant.",
        "Fabricación Nueva",
        "Merma (kg)",
        "Notas",
    ]
    col_widths = [18, 40, 20, 20, 18, 18, 18, 18, 14, 30]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # ── Data rows ──
    row = 2

    if not history_list:
        ws.cell(row=row, column=1, value="No hay modificaciones registradas")
    else:
        for entry in history_list:
            values = [
                entry.get("product_id", ""),
                entry.get("chemical_name", ""),
                entry.get("date", ""),
                entry.get("user", ""),
                entry.get("old_lote", ""),
                entry.get("new_lote", ""),
                entry.get("old_date", ""),
                entry.get("new_date", ""),
                entry.get("merma_kg", ""),
                entry.get("notes", ""),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
            row += 1

    # ── Save to temp file and return ──
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()
    wb.save(tmp_path)

    return send_file(
        tmp_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"reporte_lotes_modificados_{timestamp}.xlsx",
    )
